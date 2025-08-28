#!/usr/bin/env python3

import os
import sys
import argparse
from datetime import datetime
from pathlib import Path
import re
import logging
import traceback
import mimetypes
import csv
import stat
import tempfile
import subprocess
from typing import List, Tuple, Optional

# Enable UTF-8 output for Windows console
if sys.platform == "win32":
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer)

# Required dependencies with versions
REQUIRED_PACKAGES = {
    'PyPDF2': '3.0.1',
    'openpyxl': '3.1.2',
    'fitz': None  # PyMuPDF - optional fallback
}

def check_package_installed(package_name: str, version: str = None) -> bool:
    """Check if a package is installed and optionally check version."""
    try:
        __import__(package_name)
        return True
    except ImportError:
        return False

def install_package(package_name: str, version: str = None) -> bool:
    """Install a package using pip."""
    try:
        print(f"[INSTALLING] {package_name}...")
        package_spec = f"{package_name}=={version}" if version else package_name
        
        # Use pip module to install
        result = subprocess.run([
            sys.executable, '-m', 'pip', 'install', package_spec
        ], capture_output=True, text=True, check=True)
        
        print(f"[SUCCESS] {package_name} installed successfully")
        return True
    except subprocess.CalledProcessError as e:
        print(f"[ERROR] Failed to install {package_name}: {e}")
        print(f"[ERROR] stdout: {e.stdout}")
        print(f"[ERROR] stderr: {e.stderr}")
        return False
    except Exception as e:
        print(f"[ERROR] Unexpected error installing {package_name}: {e}")
        return False

def check_and_install_dependencies(auto_install: bool = True) -> dict:
    """Check and optionally install all required dependencies."""
    print("[INFO] Checking dependencies...")
    
    status = {
        'all_available': True,
        'installed': {},
        'missing': {},
        'install_attempted': {}
    }
    
    for package, version in REQUIRED_PACKAGES.items():
        # Special handling for PyMuPDF which imports as 'fitz'
        import_name = 'fitz' if package == 'fitz' else package
        
        if check_package_installed(import_name):
            print(f"[OK] {package} is available")
            status['installed'][package] = True
        else:
            print(f"[MISSING] {package} not found")
            status['missing'][package] = version
            status['all_available'] = False
            
            if auto_install and package != 'fitz':  # Don't auto-install PyMuPDF
                print(f"[ATTEMPTING] Auto-installing {package}...")
                success = install_package(package, version)
                status['install_attempted'][package] = success
                if success:
                    status['installed'][package] = True
                    status['all_available'] = True  # Re-check this
    
    return status

def ensure_dependencies() -> tuple:
    """Ensure dependencies are available, with user interaction if needed."""
    print("=" * 60)
    print("FILE METADATA EXTRACTOR - Dependency Check")
    print("=" * 60)
    
    # First, try to install automatically
    status = check_and_install_dependencies(auto_install=True)
    
    if status['all_available']:
        print("[SUCCESS] All dependencies are available!")
        print("=" * 60)
        return True, status
    
    # If automatic installation failed, give user options
    print("\n[WARNING] Some dependencies are still missing:")
    for package, version in status['missing'].items():
        if package not in status['install_attempted'] or not status['install_attempted'][package]:
            version_str = f"=={version}" if version else ""
            print(f"  - {package}{version_str}")
    
    print("\n[OPTIONS]")
    print("1. Continue with limited functionality (some features may not work)")
    print("2. Install missing packages manually and restart")
    print("3. Exit and install dependencies yourself")
    
    while True:
        try:
            choice = input("\nEnter your choice (1-3): ").strip()
            if choice == '1':
                print("[INFO] Continuing with available dependencies...")
                return False, status
            elif choice == '2':
                print("\n[MANUAL INSTALLATION COMMANDS]")
                for package, version in status['missing'].items():
                    if package != 'fitz':  # Skip PyMuPDF in manual instructions
                        version_str = f"=={version}" if version else ""
                        print(f"  pip install {package}{version_str}")
                print("\nAfter installing, restart the program.")
                sys.exit(0)
            elif choice == '3':
                print("[INFO] Exiting for manual dependency installation")
                sys.exit(0)
            else:
                print("[ERROR] Invalid choice. Please enter 1, 2, or 3.")
        except KeyboardInterrupt:
            print("\n[INFO] Cancelled by user")
            sys.exit(0)
        except Exception as e:
            print(f"[ERROR] Input error: {e}")
            print("[ERROR] Invalid choice. Please enter 1, 2, or 3.")

# Initialize global variables
PYPDF2_AVAILABLE = False
PYMUPDF_AVAILABLE = False  
OPENPYXL_AVAILABLE = False
PyPDF2 = None
fitz = None
Workbook = None
load_workbook = None
get_column_letter = None
Font = None
Alignment = None
Border = None
Side = None

def load_dependencies(status: dict):
    """Load dependencies based on availability status."""
    global PYPDF2_AVAILABLE, PYMUPDF_AVAILABLE, OPENPYXL_AVAILABLE
    global PyPDF2, fitz, Workbook, load_workbook, get_column_letter, Font, Alignment, Border, Side
    
    # Load PyPDF2
    if 'PyPDF2' in status['installed']:
        try:
            import PyPDF2
            PYPDF2_AVAILABLE = True
        except ImportError:
            PYPDF2_AVAILABLE = False
    
    # Load PyMuPDF (optional)
    if 'fitz' in status['installed']:
        try:
            import fitz
            PYMUPDF_AVAILABLE = True
        except ImportError:
            PYMUPDF_AVAILABLE = False
    
    # Load openpyxl
    if 'openpyxl' in status['installed']:
        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, Border, Side
            OPENPYXL_AVAILABLE = True
        except ImportError:
            OPENPYXL_AVAILABLE = False

# Setup logging
def setup_logging(debug: bool = False):
    """Setup logging configuration."""
    level = logging.DEBUG if debug else logging.INFO
    format_str = '%(asctime)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=level, format=format_str)
    return logging.getLogger(__name__)

logger = setup_logging()


def extract_number_from_filename(filename: str) -> int:
    """Extract the number from filenames like '01FileName', '02OtherFileName', etc."""
    match = re.match(r'^(\d+)', filename)
    return int(match.group(1)) if match else float('inf')


def get_file_mime_type(file_path: str) -> str:
    """Get file MIME type using multiple methods."""
    try:
        # First try mimetypes module
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type:
            logger.debug(f"MIME type for {file_path}: {mime_type}")
            return mime_type
        
        # Fallback: read file header
        with open(file_path, 'rb') as file:
            header = file.read(16)
            
        # Common file signatures
        signatures = {
            b'%PDF-': 'application/pdf',
            b'\x89PNG': 'image/png',
            b'\xff\xd8\xff': 'image/jpeg',
            b'GIF8': 'image/gif',
            b'RIFF': 'audio/wav',  # or video
            b'ID3': 'audio/mp3',
            b'\xff\xfb': 'audio/mp3',
            b'\xff\xf3': 'audio/mp3',
            b'\xff\xf2': 'audio/mp3',
            b'ftyp': 'video/mp4',  # offset 4 bytes
            b'\x1a\x45\xdf\xa3': 'video/webm',
        }
        
        for sig, mime in signatures.items():
            if header.startswith(sig):
                logger.debug(f"Detected {mime} for {file_path} by signature")
                return mime
                
        # Check MP4 at offset 4
        if len(header) >= 8 and header[4:8] == b'ftyp':
            logger.debug(f"Detected video/mp4 for {file_path} by offset signature")
            return 'video/mp4'
            
        logger.debug(f"Unknown file type for {file_path}, header: {header[:8]}")
        return 'application/octet-stream'
        
    except Exception as e:
        logger.error(f"Error detecting MIME type for {file_path}: {e}")
        return 'application/octet-stream'

def is_valid_pdf(file_path: str) -> bool:
    """Check if file is a valid PDF by reading first few bytes and MIME type."""
    try:
        mime_type = get_file_mime_type(file_path)
        is_pdf_mime = mime_type == 'application/pdf'
        
        with open(file_path, 'rb') as file:
            header = file.read(5)
            is_pdf_header = header == b'%PDF-'
            
        result = is_pdf_mime and is_pdf_header
        logger.debug(f"PDF validation for {file_path}: MIME={is_pdf_mime}, Header={is_pdf_header}, Result={result}")
        return result
        
    except Exception as e:
        logger.error(f"Error validating PDF {file_path}: {e}")
        return False

def get_pdf_pages_pypdf2(file_path: str) -> Optional[int]:
    """Get PDF pages using PyPDF2."""
    if not PYPDF2_AVAILABLE:
        logger.debug("PyPDF2 not available - skipping PyPDF2 method")
        return None
        
    try:
        logger.debug(f"Trying PyPDF2 for {file_path}")
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            page_count = len(pdf_reader.pages)
            logger.debug(f"PyPDF2 success: {page_count} pages for {file_path}")
            return page_count
    except Exception as e:
        logger.error(f"PyPDF2 failed for {file_path}: {type(e).__name__}: {str(e)}")
        logger.debug(f"PyPDF2 full traceback for {file_path}:\n{traceback.format_exc()}")
        return None

def get_pdf_pages_pymupdf(file_path: str) -> Optional[int]:
    """Get PDF pages using PyMuPDF (fallback)."""
    if not PYMUPDF_AVAILABLE:
        logger.debug("PyMuPDF not available")
        return None
    try:
        logger.debug(f"Trying PyMuPDF for {file_path}")
        doc = fitz.open(file_path)
        page_count = len(doc)
        doc.close()
        logger.debug(f"PyMuPDF success: {page_count} pages for {file_path}")
        return page_count
    except Exception as e:
        logger.error(f"PyMuPDF failed for {file_path}: {type(e).__name__}: {str(e)}")
        logger.debug(f"PyMuPDF full traceback for {file_path}:\n{traceback.format_exc()}")
        return None

def get_pdf_pages_estimate(file_path: str) -> int:
    """Estimate PDF pages by searching for page objects in raw content."""
    try:
        logger.debug(f"Trying estimation method for {file_path}")
        with open(file_path, 'rb') as file:
            content = file.read(min(1024*1024, os.path.getsize(file_path)))  # Read max 1MB
            
        # Count occurrences of page object patterns
        page_patterns = [b'/Type /Page', b'/Type/Page', b'endobj']
        pattern_counts = {}
        
        for pattern in page_patterns:
            count = content.count(pattern)
            if pattern == b'endobj':
                count = max(1, count // 10)  # Rough estimate
            pattern_counts[pattern] = count
            
        max_count = max(pattern_counts.values()) if pattern_counts else 1
        estimated_pages = max(1, min(max_count, 1000))  # Cap at reasonable number
        
        logger.debug(f"Pattern counts for {file_path}: {pattern_counts}")
        logger.debug(f"Estimated {estimated_pages} pages for {file_path}")
        return estimated_pages
        
    except Exception as e:
        logger.error(f"Estimation failed for {file_path}: {type(e).__name__}: {str(e)}")
        return 1

def should_process_as_pdf(file_path: str) -> bool:
    """Determine if file should be processed as PDF based on extension AND content."""
    # Only check files with .pdf extension
    if not file_path.lower().endswith('.pdf'):
        return False
        
    # Check if it's actually a PDF file
    mime_type = get_file_mime_type(file_path)
    is_actually_pdf = is_valid_pdf(file_path)
    
    logger.debug(f"PDF check for {file_path}: extension=.pdf, mime={mime_type}, valid_pdf={is_actually_pdf}")
    
    if not is_actually_pdf:
        logger.warning(f"File {file_path} has .pdf extension but is not a valid PDF (MIME: {mime_type})")
        return False
        
    return True

def get_pdf_pages(file_path: str) -> int:
    """Get number of pages in a PDF file with comprehensive error handling."""
    logger.info(f"Processing PDF: {file_path}")
    
    # First check if we should even try to process as PDF
    if not should_process_as_pdf(file_path):
        logger.warning(f"Skipping PDF processing for {file_path} - not a valid PDF")
        return 1
    
    # Try PyPDF2 first
    try:
        pages = get_pdf_pages_pypdf2(file_path)
        if pages is not None:
            logger.info(f"Successfully extracted {pages} pages from {file_path} using PyPDF2")
            return pages
    except Exception as e:
        logger.error(f"Unexpected error in PyPDF2 processing for {file_path}: {e}")
    
    # Try PyMuPDF as fallback
    try:
        pages = get_pdf_pages_pymupdf(file_path)
        if pages is not None:
            logger.info(f"Successfully extracted {pages} pages from {file_path} using PyMuPDF (fallback)")
            return pages
    except Exception as e:
        logger.error(f"Unexpected error in PyMuPDF processing for {file_path}: {e}")
    
    # Last resort: estimate based on file content
    try:
        pages = get_pdf_pages_estimate(file_path)
        logger.warning(f"Using estimation method for {file_path}: {pages} pages")
        return pages
    except Exception as e:
        logger.error(f"All PDF processing methods failed for {file_path}: {e}")
        logger.error(f"Final fallback traceback:\n{traceback.format_exc()}")
        return 1


def get_file_size(file_path: str) -> str:
    """Get file size formatted as KB or MB."""
    try:
        size_bytes = os.path.getsize(file_path)
        if size_bytes < 1024:
            return f"{size_bytes} bytes"
        elif size_bytes < 1024 * 1024:
            size_kb = size_bytes / 1024
            return f"{size_kb:.1f} KB".replace('.', ',')
        else:
            size_mb = size_bytes / (1024 * 1024)
            return f"{size_mb:.2f} MB".replace('.', ',')
    except Exception:
        return "0 bytes"


def count_files_in_directory(dir_path: str) -> int:
    """Count number of files in a directory."""
    try:
        return len([f for f in os.listdir(dir_path) if os.path.isfile(os.path.join(dir_path, f))])
    except Exception:
        return 0


def get_creation_date(file_path: str) -> str:
    """Get file creation date formatted as Spanish date."""
    try:
        stat = os.stat(file_path)
        # Use birth time if available (macOS), otherwise use modification time
        creation_time = getattr(stat, 'st_birthtime', stat.st_mtime)
        dt = datetime.fromtimestamp(creation_time)
        # Format as Spanish date: d/mm/yyyy h:mm a. m./p. m.
        am_pm = "a. m." if dt.hour < 12 else "p. m."
        hour_12 = dt.hour if dt.hour <= 12 else dt.hour - 12
        if hour_12 == 0:
            hour_12 = 12
        return f"{dt.day}/{dt.month:02d}/{dt.year} {hour_12}:{dt.minute:02d} {am_pm}"
    except Exception:
        dt = datetime.now()
        am_pm = "a. m." if dt.hour < 12 else "p. m."
        hour_12 = dt.hour if dt.hour <= 12 else dt.hour - 12
        if hour_12 == 0:
            hour_12 = 12
        return f"{dt.day}/{dt.month:02d}/{dt.year} {hour_12}:{dt.minute:02d} {am_pm}"


def get_ordered_files(directory: str) -> List[Tuple[str, str]]:
    """Get files and directories ordered by their numeric prefix."""
    items = []
    
    for item in os.listdir(directory):
        item_path = os.path.join(directory, item)
        items.append((item, item_path))
    
    # Sort by the numeric prefix extracted from filename
    items.sort(key=lambda x: extract_number_from_filename(x[0]))
    
    return items


def setup_excel_formatting():
    """Setup Excel formatting styles."""
    # Header font and style
    header_font = Font(name='Calibri', size=11, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data font and style
    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(vertical='center', wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return header_font, header_alignment, data_font, data_alignment, thin_border

def add_headers_and_formatting(ws, directory: str = "", litigant_name: str = ""):
    """Add headers and formatting to match the template."""
    header_font, header_alignment, data_font, data_alignment, thin_border = setup_excel_formatting()
    
    # Add document header information to match indice de ejemplo format
    # Row 1: Ciudad
    ws['A1'].value = 'Ciudad'
    ws['B1'].value = 'SANTIAGO DE CALI (VALLE)'
    ws['H1'].value = 'EXPEDIENTE FÍSICO'
    
    # Row 2: Despacho Judicial (Judge)
    ws['A2'].value = 'Despacho Judicial'
    ws['B2'].value = 'JUZGADO DECIMO LABORALDEL  CIRCUITO DE CALI'  # Standard judge format
    ws['H2'].value = 'El expediente judicial posee documentos físicos:'
    ws['J2'].value = 'SI     NO X'
    
    # Row 3: Serie
    ws['A3'].value = 'Serie o Subserie Documental'
    ws['B3'].value = 'ORDINARIO LABORAL DE PRIMERA INSTANCIA'
    
    # Row 4: Radicación  
    ws['A4'].value = 'No. Radicación del Proceso'
    ws['B4'].value = '76001310501020230040100'  # Standard format, could be made configurable
    ws['H4'].value = 'No. de carpetas (cuadernos), legajos o tomos:'
    
    # Row 5: Demandado
    ws['A5'].value = 'Partes Procesales (Parte A)\n(demandado, procesado, accionado)'
    ws['B5'].value = 'PORVENIR Y OTROS'  # Standard format
    ws['H5'].value = 'No. de carpetas (cuadernos), legajos o tomos digitalizados:'
    
    # Row 6: Demandante (Litigant)
    ws['A6'].value = 'Partes Procesales (Parte B)\n(demandante, denunciante, accionante)'
    ws['B6'].value = litigant_name.upper() if litigant_name else 'NOMBRE DEL LITIGANTE'
    
    # Row 7: Terceros
    ws['A7'].value = 'Terceros Intervinientes'
    
    # Row 8: Cuaderno
    ws['A8'].value = 'Cuaderno '
    
    # Row 10: Main title
    ws['A10'].value = 'ÍNDICE ELECTRÓNICO DEL EXPEDIENTE JUDICIAL'
    
    # Headers in Spanish (row 11)
    headers = {
        'A11': 'Nombre Documento',
        'B11': 'Fecha Creación Documento', 
        'C11': 'Fecha Incorporación Expediente',
        'D11': 'Orden Documento',
        'E11': 'Número Páginas',
        'F11': 'Página Inicio',
        'G11': 'Página Fin',
        'H11': 'Formato',
        'I11': 'Tamaño',
        'J11': 'Origen',
        'K11': 'Observaciones'
    }
    
    # Add headers with formatting
    for cell_ref, header_text in headers.items():
        cell = ws[cell_ref]
        cell.value = header_text
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

def check_file_permissions(file_path: str) -> bool:
    """Check if we can write to the specified file path."""
    try:
        # Check if file exists and is writable
        if os.path.exists(file_path):
            return os.access(file_path, os.W_OK)
        
        # Check if directory is writable
        directory = os.path.dirname(file_path) or '.'
        return os.access(directory, os.W_OK)
    except Exception as e:
        logger.error(f"Error checking permissions for {file_path}: {e}")
        return False

def create_safe_filename(base_name: str) -> str:
    """Create a safe filename, adding timestamp if needed."""
    if check_file_permissions(base_name):
        return base_name
    
    # Try with timestamp
    name, ext = os.path.splitext(base_name)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    timestamped_name = f"{name}_{timestamp}{ext}"
    
    if check_file_permissions(timestamped_name):
        logger.info(f"Using timestamped filename: {timestamped_name}")
        return timestamped_name
    
    # Fall back to temp directory
    temp_dir = tempfile.gettempdir()
    temp_name = os.path.join(temp_dir, os.path.basename(timestamped_name))
    logger.warning(f"Using temp directory: {temp_name}")
    return temp_name

def export_to_csv(data: List[dict], csv_file: str) -> bool:
    """Export data to CSV format."""
    try:
        logger.info(f"Exporting to CSV: {csv_file}")
        
        # Create safe filename
        safe_csv_file = create_safe_filename(csv_file)
        
        with open(safe_csv_file, 'w', newline='', encoding='utf-8') as file:
            if not data:
                logger.warning("No data to export to CSV")
                return False
                
            writer = csv.DictWriter(file, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            
        logger.info(f"CSV export successful: {safe_csv_file}")
        print(f"CSV file created: {safe_csv_file}")
        return True
        
    except Exception as e:
        logger.error(f"CSV export failed: {e}")
        logger.error(f"CSV export traceback:\n{traceback.format_exc()}")
        return False

def process_files_to_excel(directory: str, output_file: str = "metadata_output.xlsx", export_csv: bool = False, litigant_name: str = ""):
    """Process files and write metadata to Excel file matching the template format."""
    
    # Check if we can create Excel files
    if not OPENPYXL_AVAILABLE and not export_csv and not output_file.endswith('.csv'):
        logger.error("Cannot create Excel files - openpyxl not available. Switching to CSV output.")
        print("[WARNING] Switching to CSV output due to missing dependencies")
        output_file = output_file.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
        export_csv = True
    
    # Get ordered files
    ordered_items = get_ordered_files(directory)
    
    # Only create Excel workbook if openpyxl is available and we need Excel output
    wb = None
    ws = None
    if OPENPYXL_AVAILABLE and not output_file.endswith('.csv'):
        # Create new workbook (always start fresh to match template)
        wb = Workbook()
        ws = wb.active
        ws.title = "Indice Electrónico"
    
    # Add headers and formatting (only for Excel)
    if ws is not None:
        add_headers_and_formatting(ws, directory, litigant_name)
    
    # Setup Excel formatting (only if needed)
    header_font, header_alignment, data_font, data_alignment, thin_border = None, None, None, None, None
    if ws is not None:
        header_font, header_alignment, data_font, data_alignment, thin_border = setup_excel_formatting()
    
    # Starting row for data (A12 as specified)
    start_row = 12
    current_page = 1  # Track page numbering
    
    # Collect all data first (works with or without Excel)
    data_rows = []
    
    for idx, (item_name, item_path) in enumerate(ordered_items):
        current_row = start_row + idx
        
        # Collect data for this item
        creation_date_str = get_creation_date(item_path)
        row_data = {
            'Nombre Documento': item_name,
            'Fecha Creación Documento': creation_date_str,
            'Fecha Incorporación Expediente': creation_date_str,
            'Orden Documento': idx + 1,
            'Página Inicio': current_page,
        }
        
        # Write to Excel if available
        if ws is not None:
            ws[f'A{current_row}'].value = item_name
            ws[f'B{current_row}'].value = creation_date_str
            ws[f'C{current_row}'].value = creation_date_str
            ws[f'D{current_row}'].value = idx + 1
        
        if os.path.isfile(item_path):
            # Process pages
            try:
                if item_path.lower().endswith('.pdf'):
                    logger.info(f"Processing potential PDF: {item_path}")
                    pages = get_pdf_pages(item_path)
                else:
                    pages = 1
                    logger.debug(f"Non-PDF file: {item_path}, setting pages = 1")
            except Exception as e:
                logger.error(f"Error processing file {item_path} for page count: {e}")
                logger.error(f"Stack trace:\n{traceback.format_exc()}")
                pages = 1  # Fallback
            
            # File format and size
            file_ext = os.path.splitext(item_path)[1].upper().lstrip('.')
            file_size = get_file_size(item_path)
            
            # Update row data
            row_data.update({
                'Número Páginas': pages,
                'Página Fin': current_page + pages - 1,
                'Formato': file_ext if file_ext else "UNKNOWN",
                'Tamaño': file_size,
                'Origen': "ELECTRONICO",
                'Observaciones': ""
            })
            
            # Write to Excel if available
            if ws is not None:
                ws[f'E{current_row}'].value = pages
                ws[f'F{current_row}'].value = current_page
                ws[f'G{current_row}'].value = f"=F{current_row}+E{current_row}-1"
                ws[f'H{current_row}'].value = file_ext if file_ext else "UNKNOWN"
                ws[f'I{current_row}'].value = file_size
                ws[f'J{current_row}'].value = "ELECTRONICO"
            
            # Update page counter for next file
            current_page += pages
            
        elif os.path.isdir(item_path):
            # For directories
            file_count = count_files_in_directory(item_path)
            
            # Update row data
            row_data.update({
                'Número Páginas': 1,
                'Página Fin': current_page,
                'Formato': "CARPETA",
                'Tamaño': f"{file_count} archivos",
                'Origen': "ELECTRONICO",
                'Observaciones': ""
            })
            
            # Write to Excel if available
            if ws is not None:
                ws[f'E{current_row}'].value = 1  # Directories count as 1 page
                ws[f'F{current_row}'].value = current_page
                ws[f'G{current_row}'].value = current_page  # Same start and end for directories
                ws[f'H{current_row}'].value = "CARPETA"
                ws[f'I{current_row}'].value = f"{file_count} archivos"
                ws[f'J{current_row}'].value = "ELECTRONICO"
            
            current_page += 1
        
        # Add row to data collection
        data_rows.append(row_data)
        
        # Apply formatting to all data cells (Excel only)
        if ws is not None and header_font is not None:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                cell = ws[f'{col}{current_row}']
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
    
    # Adjust column widths (Excel only)
    if ws is not None:
        column_widths = {
            'A': 25,  # Nombre Documento
            'B': 18,  # Fecha Creación
            'C': 18,  # Fecha Incorporación  
            'D': 8,   # Orden
            'E': 10,  # Número Páginas
            'F': 10,  # Página Inicio
            'G': 10,  # Página Fin
            'H': 12,  # Formato
            'I': 15,  # Tamaño
            'J': 12,  # Origen
            'K': 20   # Observaciones
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    # Use collected data for CSV export (works with or without Excel)
    csv_data = data_rows if export_csv or output_file.endswith('.csv') else []
    
    # Save the Excel workbook (only if Excel is available and needed)
    if wb is not None and not output_file.endswith('.csv'):
        try:
            safe_excel_file = create_safe_filename(output_file)
            logger.info(f"Saving Excel file: {safe_excel_file}")
            wb.save(safe_excel_file)
            print(f"Metadata extracted and saved to {safe_excel_file}")
            print(f"Processed {len(ordered_items)} items from {directory}")
            print(f"Format matches 'indice de ejemplo.xlsm' template")
        except PermissionError as e:
            logger.error(f"Permission denied when saving Excel file: {e}")
            print(f"ERROR: Permission denied when saving {output_file}")
            print("Possible causes:")
            print("- File is open in Excel or another program")
            print("- Insufficient write permissions in the directory")
            print("Trying CSV export instead...")
            csv_file = output_file.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
            export_to_csv(csv_data, csv_file)
            return
        except Exception as e:
            logger.error(f"Unexpected error saving Excel file: {e}")
            logger.error(f"Full traceback:\n{traceback.format_exc()}")
            print(f"ERROR: Failed to save Excel file: {e}")
            print("Trying CSV export instead...")
            csv_file = output_file.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
            export_to_csv(csv_data, csv_file)
            return
        
        # Export to CSV if requested
        if export_csv:
            csv_file = safe_excel_file.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
            export_to_csv(csv_data, csv_file)
    
    # Handle CSV-only output or fallback to CSV
    elif csv_data:
        csv_file = create_safe_filename(output_file) if output_file.endswith('.csv') else output_file.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
        export_to_csv(csv_data, csv_file)
        print(f"Metadata extracted and saved to {csv_file}")
        print(f"Processed {len(ordered_items)} items from {directory}")
        if not OPENPYXL_AVAILABLE:
            print("[INFO] CSV format used due to missing Excel dependencies")
        print(f"Format matches 'indice de ejemplo.xlsm' template")


def main():
    # Check and install dependencies FIRST, before any user interaction
    deps_ok, dep_status = ensure_dependencies()
    
    # Load the available dependencies  
    load_dependencies(dep_status)
    
    parser = argparse.ArgumentParser(description='Extract file metadata and write to Excel')
    parser.add_argument('--directory', '-d', 
                       default=os.getcwd(),
                       help='Directory to process (default: current directory)')
    parser.add_argument('--output', '-o',
                       default='metadata_output.xlsx',
                       help='Output Excel file (default: metadata_output.xlsx)')
    parser.add_argument('--debug', '-v',
                       action='store_true',
                       help='Enable debug logging')
    parser.add_argument('--log-file',
                       help='Write logs to file instead of console')
    parser.add_argument('--csv', '-c',
                       action='store_true',
                       help='Also export to CSV format')
    parser.add_argument('--csv-only',
                       action='store_true',
                       help='Export only to CSV format (no Excel)')
    parser.add_argument('--litigant', '-l',
                       default='',
                       help='Name of the litigant (demandante) for the header')
    
    args = parser.parse_args()
    
    # Setup logging based on arguments
    global logger
    if args.log_file:
        # Setup file logging
        file_handler = logging.FileHandler(args.log_file)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
    if args.debug:
        logger.setLevel(logging.DEBUG)
        logging.getLogger().setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")
    
    logger.info(f"Starting file metadata extraction from: {args.directory}")
    logger.info(f"Output file: {args.output}")
    logger.info(f"Debug mode: {args.debug}")
    
    if not os.path.isdir(args.directory):
        logger.error(f"Directory not found: {args.directory}")
        print(f"Error: {args.directory} is not a valid directory")
        sys.exit(1)
    
    try:
        if args.csv_only:
            # CSV-only mode
            csv_file = args.output.replace('.xlsx', '.csv').replace('.xlsm', '.csv')
            logger.info("CSV-only mode enabled")
            process_files_to_excel(args.directory, csv_file, export_csv=True, litigant_name=args.litigant)
        else:
            # Normal mode (Excel + optional CSV)
            process_files_to_excel(args.directory, args.output, export_csv=args.csv, litigant_name=args.litigant)
            
        logger.info("Processing completed successfully")
        
    except PermissionError as e:
        logger.error(f"Permission error: {e}")
        print(f"\nERROR: Permission denied - {e}")
        print("\nTroubleshooting steps:")
        print("1. Close Excel or any program that might have the file open")
        print("2. Check file permissions and make sure directory is writable")
        print("3. Try running with --csv-only flag for CSV output only")
        print("4. Try specifying a different output directory with -o")
        sys.exit(1)
        
    except Exception as e:
        logger.error(f"Fatal error during processing: {e}")
        logger.error(f"Full traceback:\n{traceback.format_exc()}")
        print(f"\nERROR: {e}")
        if args.debug:
            print("\nCheck debug logs for detailed error information.")
        print("\nTry using --csv-only flag if Excel export is failing.")
        sys.exit(1)


if __name__ == "__main__":
    main()